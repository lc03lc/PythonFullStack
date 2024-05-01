import re
import json
import os
import time
import datetime

from openpyxl import load_workbook
from utils import req
from config import settings


class PanHandler(object):
    def __init__(self, conn):
        # 客户端的socket连接
        self.conn = conn
        self.userName = None

    @property
    def home_path(self):
        return os.path.join(settings.USER_FOLD_PATH, self.userName)

    def send_json_data(self, **kwargs):
        req.send_data(self.conn, json.dumps(kwargs))

    def recv_save_file(self, target_file_path):
        req.recv_save_file(self.conn, target_file_path)

    def send_file_by_seek(self, file_size, file_path, seek=0):
        req.send_file_by_seek(self.conn, file_size, file_path, seek)

    def login(self, username, password):
        '''用户登录：读取excel文件，进行用户登录'''
        wb = load_workbook(settings.DB_FILE_PATH)
        sheet = wb.worksheets[0]

        success = False
        for row in sheet.iter_rows(2):
            if username == row[0].value and password == row[1].value:
                success = True
                break

        if success:
            self.send_json_data(status=True, data="登录成功")
            self.userName = username
        else:
            self.send_json_data(status=False, error="登录失败")

    def register(self, username, password):
        '''用户注册：用户名和密码写入excel中'''
        wb = load_workbook(settings.DB_FILE_PATH)
        sheet = wb.worksheets[0]
        # 检测用户是否存在
        exists = False
        for row in sheet.iter_rows(2):
            if username == row[0].value:
                exists = True
                break
        if exists:
            # 给用户回复已存在
            self.send_json_data(status=False, error="用户名已存在")
            return
        # 注册写入excel
        max_row = sheet.max_row
        data_list = [username, password, datetime.now().strftime("%Y-%m-%d")]
        for i, item in enumerate(data_list, 1):
            cell = sheet.cell(max_row + 1, i)
            cell.value = item
        wb.save(settings.DB_FILE_PATH)

        # 创建目录
        user_folder = os.path.join(settings.USER_FOLD_PATH, username)
        os.makedirs(user_folder)

        # 回复消息
        self.send_json_data(status=True, data="注册成功")

    def ls(self, folder_path=None):
        '''
        查看用户目录下的文件
        如果folder_path=None,查看用户根目录
        如果folder_path不为空，查看目录/folder_path中的文件
        '''
        if not self.userName:
            self.send_json_data(status=False, error="登录后才能查看")
            return

        if not folder_path:
            data = "\n".join(os.listdir(self.home_path))
            self.send_json_data(status=True, data=data)
            return

        target_folder_path = os.path.join(self.home_path, folder_path)
        if not os.path.exists(target_folder_path):
            self.send_json_data(status=False, error="路径不存在")
            return
        if not os.path.isdir(target_folder_path):
            self.send_json_data(status=False, error="文件夹不存在")

        data = '\n'.join(os.listdir(target_folder_path))
        self.send_json_data(status=True, data=data)

    def upload(self, file_path):
        '''上传文件，直接覆盖'''
        if not self.userName:
            self.send_json_data(status=False, error="登录后才能上传")
            return

        target_file_path = os.path.join(self.home_path, file_path)
        folder = os.path.dirname(target_file_path)
        if not os.path.exists(folder):
            os.mkdir(folder)

        self.send_json_data(stastus=True, data='开始上传')
        time.sleep(2)
        # 很关键，如果是基于select+非阻塞模式，直接去recv，会出现blocking error错误

        self.recv_save_file(target_file_path)

    def download(self, file_path, seek=0):
        '''下载文件，支持断点续传'''
        if not self.userName:
            self.send_json_data(status=False, error="登录后才能下载")
            return

        # 文件不存在
        target_file_path = os.path.join(self.home_path, file_path)
        if not os.path.exists(target_file_path):
            self.send_json_data(status=False, error="文件{}不存在".format(file_path))
            return

        # 获取文件大小并返回
        self.send_json_data(status=True, data="开始下载")

        # 发送文件
        seek = int(seek)
        total_size = os.stat(target_file_path).st_size
        req.send_file_by_seek(self.conn, total_size-seek, target_file_path, seek)

    def execute(self):
        '''
        每次客户端发来请求，触发此方法
        :return: False，关闭连接；True，继续连接请求
        '''
        conn = self.conn

        # 登录/注册/查看目录/上传文件/下载文件
        # 1. 获取数据包
        cmd = req.recv_data(conn).decode('utf-8')
        if cmd.upper() == 'Q':
            print("客户端退出")
            return False
        method_map = {
            'login': self.login,
            'register': self.register,
            'ls': self.ls,
            'upload': self.upload,
            'download': self.download
        }
        cmd, *args = re.split(r"\s+", cmd)
        method = method_map[cmd]

        # 2.执行相关方法
        method(*args)

        return True
